#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: ai ts=4 sts=4 et sw=4 nu

from __future__ import (unicode_literals, absolute_import,
                        division, print_function)
import os
import random
import datetime

from django.utils import timezone
from optparse import make_option
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from douentza.models import HotlineRequest
from douentza.helpers import create_request
from douentza.utils import normalize_phone_number

NB_DAYS = 7


def random_date(start, end):
    delta = end - start
    int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
    random_second = random.randrange(int_delta)
    return start + datetime.timedelta(seconds=random_second)


class Command(BaseCommand):
    help = "Create HotlineRequest from Excel list of phone numbers"

    option_list = BaseCommand.option_list + (
        make_option('-f',
                    help='Excel file to import numbers from',
                    action='store',
                    dest='input_file'),
        )

    def handle(self, *args, **options):

        filepath = options.get('input_file')

        if filepath is None or not os.path.exists(filepath):
            print("Unable to open Excel file: `{}`".format(filepath))
            return

        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        nb_calls = len(ws.rows)
        cd = lambda row, column: ws.cell(row=row, column=column).value

        end = timezone.now()
        start = end - datetime.timedelta(days=NB_DAYS)
        step = NB_DAYS * 86400 / nb_calls
        received_on = start

        for row in range(1, nb_calls):
            value = cd(row, 1)
            received_on += datetime.timedelta(seconds=step)
            try:
                value = int(value.strip())
            except:
                if not isinstance(value, int):
                    continue
            print(value)

            identity = normalize_phone_number(str(value))
            req = create_request(
                identity=identity,
                event_type=HotlineRequest.TYPE_RING,
                received_on=received_on,
                operator=None,
                message=None,
                phone_number=None)

            print(req)
