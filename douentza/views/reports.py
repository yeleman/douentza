#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: ai ts=4 sts=4 et sw=4 nu

from __future__ import (unicode_literals, absolute_import,
                        division, print_function)
import datetime
import copy
import io

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl.styles import (PatternFill, Border, Side,
                             Alignment, Protection, Font)
from openpyxl.styles.fills import FILL_SOLID
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

from douentza.models import Cluster, HotlineRequest


def letter_to_column(letter):
    return openpyxl.cell.column_index_from_string(letter)


def column_to_letter(column):
    return openpyxl.cell.get_column_letter(column)


def xl_col_width(cm):
    """ xlwt width for a given width in centimeters """
    return 12 * cm


def xl_set_col_width(sheet, column, cm):
    """ change column width """
    letter = column_to_letter(column)
    if column not in sheet.column_dimensions.keys():
        sheet.column_dimensions[letter] = \
            ColumnDimension(worksheet=sheet)
    sheet.column_dimensions[letter].width = xl_col_width(cm)


def xl_row_height(cm):
    """ xlwt height for a given height in centimeters """
    return 100 * cm


def xl_set_row_height(sheet, row, cm):
    """ change row height """
    if row not in sheet.row_dimensions.keys():
        sheet.row_dimensions[row] = RowDimension(worksheet=sheet)
    sheet.row_dimensions[row].height = xl_row_height(cm)


dataentry_fname_for = lambda cluster, weeknum: "logbook-{}-W{}.xlsx".format(
    cluster.slug, weeknum)

# colors
black = 'FF000000'
dark_gray = 'FFA6A6A6'
light_gray = 'FFDEDEDE'
yellow = 'F9FF00'
green = '2C8015'
saumon = '864666'

# styles
header_font = Font(
    name='Calibri',
    size=12,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color=black)

mheader_font = copy.copy(header_font)
mheader_font.size = 14

std_font = Font(
    name='Calibri',
    size=12,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color=black)

header_fill = PatternFill(fill_type=FILL_SOLID, start_color=dark_gray)
day_header_fill = PatternFill(fill_type=FILL_SOLID, start_color=green)
week_header_fill = PatternFill(fill_type=FILL_SOLID, start_color=saumon)

thin_black_side = Side(style='thin', color='FF000000')
thick_black_side = Side(style='thick', color='FF000000')
no_border = Side(style='none')

std_border = Border(
    left=thin_black_side,
    right=thin_black_side,
    top=thin_black_side,
    bottom=thin_black_side,
)

thick_left_border = Border(
    left=thick_black_side,
    right=thin_black_side,
    top=thin_black_side,
    bottom=thin_black_side,)
thick_right_border = Border(
    right=thick_black_side,
    left=thin_black_side,
    top=thin_black_side,
    bottom=thin_black_side,)
thick_top_border = Border(
    right=thick_black_side,
    left=thick_black_side,
    top=thick_black_side,
    bottom=no_border)
thick_bottom_border = Border(
    right=thick_black_side,
    left=thick_black_side,
    top=no_border,
    bottom=thick_black_side)

centered_alignment = Alignment(
    horizontal='center',
    vertical='center',
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0)

left_alignment = Alignment(
    horizontal='left',
    vertical='top')

right_alignment = Alignment(
    horizontal='right',
    vertical='top')

left_wrapped_alignment = Alignment(
    horizontal='left',
    vertical='top',
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0)

number_format = '# ### ### ##0'

protected = Protection(locked=True, hidden=False)
unprotected = Protection(locked=False, hidden=False)

header_style = {
    'font': header_font,
    'fill': header_fill,
    'border': std_border,
    'alignment': left_alignment,
    'protection': protected
}

mheader_style = copy.copy(header_style)
mheader_style.update({'alignment': centered_alignment,
                      'font': mheader_font})
mheadertop_style = copy.copy(mheader_style)
mheadertop_style.update({'border': thick_top_border})
mheaderbottom_style = copy.copy(mheader_style)
mheaderbottom_style.update({'border': thick_bottom_border})

dheader_style = copy.copy(header_style)
dheader_style.update({'fill': day_header_fill})

wheader_style = copy.copy(header_style)
wheader_style.update({'fill': week_header_fill})

std_style = {
    'font': std_font,
    'border': std_border,
    'alignment': left_alignment,
}

value_style = copy.copy(std_style)
value_style.update({'alignment': right_alignment})

stdw_style = copy.copy(std_style)
stdw_style.update({'alignment': left_wrapped_alignment})

bold_style = copy.copy(std_style)
bold_style.update({'font': header_font})

names_style = {
    'font': std_font,
    'border': std_border,
    'alignment': left_alignment,
}


def apply_style(target, style):
    for key, value in style.items():
        setattr(target, key, value)


def eom(adate):
    ndate = adate + datetime.timedelta(days=31)
    return datetime.datetime(ndate.year, ndate.month, 1, 23, 59, 59,
                             tzinfo=datetime.timezone.utc) \
        - datetime.timedelta(days=1)


def eow(adate):
    ndate = adate + datetime.timedelta(days=7)
    return datetime.datetime(ndate.year, ndate.month, ndate.day, 23, 59, 59,
                             tzinfo=datetime.timezone.utc) \
        - datetime.timedelta(days=1)


def day_bound(aday):
    return (datetime.datetime(aday.year, aday.month, aday.day, 0, 0, 1,
                              tzinfo=datetime.timezone.utc),
            datetime.datetime(aday.year, aday.month, aday.day, 23, 59, 59,
                              tzinfo=datetime.timezone.utc))


def week_bound(monday):
    sunday = monday + datetime.timedelta(days=7)
    return (datetime.datetime(monday.year, monday.month, monday.day, 0, 0, 1,
                              tzinfo=datetime.timezone.utc),
            datetime.datetime(sunday.year, sunday.month, sunday.day,
                              23, 59, 59, tzinfo=datetime.timezone.utc))


def days_for_period(start, end):
    days = []
    assert start < end
    day = start.date()
    while day < end.date():
        days.append(day_bound(day))
        day = day + datetime.timedelta(days=1)
    days.append(day_bound(day))
    return days


def weeks_for_period(start, end):
    weeks = []
    monday = start.date() - datetime.timedelta(
        days=start.date().isoweekday() - 1)
    while monday < end.date():
        m, s = week_bound(monday)
        weeks.append((monday.strftime("%Y-W%W"), m, s))
        monday = s.date() + datetime.timedelta(days=1)
    return weeks


def gen_report(lang, period):
    cluster = Cluster.get_or_none(lang)
    assert cluster is not None
    assert period is not None

    # p_year, p_month = [int(x) for x in period.split("-")]
    # start_date = datetime.datetime(p_year, p_month, 1, 0, 0, 1)
    start_date = (
        datetime.datetime.strptime(period + '-0', "%Y-W%W-%w") +
        datetime.timedelta(days=1)).replace(tzinfo=datetime.timezone.utc)
    end_date = eow(start_date)
    assert end_date > start_date

    hr_requests = HotlineRequest.done \
        .filter(cluster__slug=cluster.slug) \
        .filter(
            received_on__gte=start_date,
            received_on__lte=end_date) \
        .order_by('received_on')

    days = days_for_period(start_date, end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "LOGBOOK"

    reasons = ["Appreciation", "Agriculture", "Corruption", "Health",
               "Security", "Complain", "Hunger", "Request"]
    reasons_validation = DataValidation(
        type="list", formula1='"{}"'.format(",".join(reasons)),
        allow_blank=True)
    ws.add_data_validation(reasons_validation)

    def std_write(row, column, value, style=std_style):
        cell = ws.cell(row=row, column=column)
        cell.value = value
        apply_style(cell, style)

    headers = ["S/N", "Date", "Transcript", "Issue",
               "Location", "Ward", "LGA", "State", "Summary of Issue",
               "Remarks"]
    sn_col, date_col, transcript_col, issue_col, location_col, ward_col, \
        lga_col, state_col, summary_col, remarks_col = \
        range(1, len(headers) + 1)

    state_letter = column_to_letter(state_col)
    issue_letter = column_to_letter(issue_col)

    day_headers = ["", "", "Summary for the day", "", "",
                   "Adamawa", "Borno", "Yobe", "Incoming calls"]
    sn_dcol, date_dcol, summary_dcol, _, _, adamawa_dcol, \
        borno_dcol, yobe_dcol, total_dcol = range(1, len(day_headers) + 1)

    borno_letter = column_to_letter(borno_dcol)
    adamawa_letter = column_to_letter(adamawa_dcol)
    yobe_letter = column_to_letter(yobe_dcol)

    def style_row(row, style):
        for column in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=column)
            apply_style(cell, style)

    def apply_first_cols_style(row, colnum, style):
        for column in range(1, colnum):
            cell = ws.cell(row=row, column=column)
            apply_style(cell, style)

    # top header with period and lang
    ws.merge_cells("A1:{ll}1".format(ll=column_to_letter(len(headers))))
    std_write(1, 1, "DIGITAL DEVELOPMENT HUB (DD-HUB)", mheadertop_style)
    style_row(1, mheadertop_style)
    ws.merge_cells("A2:{ll}2".format(ll=column_to_letter(len(headers))))
    std_write(2, 1, "DD-HUB-NERI DAILY LOG BOOK/{lang}/{period}"
                    .format(lang=cluster.name, period=period),
                    mheaderbottom_style)
    style_row(2, mheaderbottom_style)

    # write top header with indic name
    for colnum, header in enumerate(headers):
        std_write(3, colnum + 1, header, header_style)

    # column widths
    xl_set_col_width(ws, sn_col, 0.45)
    xl_set_col_width(ws, transcript_col, 3.1)
    xl_set_col_width(ws, issue_col, 1.6)
    xl_set_col_width(ws, location_col, 1.6)
    xl_set_col_width(ws, state_col, 1.2)
    xl_set_col_width(ws, summary_col, 2.2)
    xl_set_col_width(ws, remarks_col, 1.8)

    data_first_row = 4
    row = copy.copy(data_first_row)
    for num, day in enumerate(days):

        day_requests = hr_requests.filter(
            received_on__gte=day[0],
            received_on__lte=day[1]).order_by('received_on')

        has_data = day_requests.count() > 0

        day_start = copy.copy(row)

        # if no request for given day
        if not has_data:
            continue
            style_row(row, std_style)
            # blank line with the date
            std_write(row, date_col,
                      day[0].date().strftime("%Y-%m-%d"), std_style)
            ws.merge_cells("{ca}{r}:{cb}{r}".format(
                r=row,
                ca=column_to_letter(transcript_col),
                cb=column_to_letter(len(headers) + 1)))
            row += 1
            continue

        # one line per request
        for reqnum, req in enumerate(day_requests.all()):
            # style row
            xl_set_row_height(ws, row, 0.7)
            style_row(row, std_style)

            reasons_validation.ranges.append('{l}{r}'.format(
                r=row, l=column_to_letter(issue_col)))

            # meta data
            std_write(row, sn_col, req.id, std_style)
            std_write(row, date_col,
                      day[0].date().strftime("%Y-%m-%d"), std_style)

            std_write(row, transcript_col, req.transcript, stdw_style)

            if req.location:
                std_write(row, location_col, str(req.location), std_style)
                if req.location.get_ward():
                    std_write(row, ward_col,
                              str(req.location.get_ward()), std_style)
                if req.location.get_lga():
                    std_write(row, lga_col,
                              str(req.location.get_lga()), std_style)
                if req.location.get_state():
                    std_write(row, state_col,
                              str(req.location.get_state()), std_style)
            row += 1

        day_end = copy.copy(row - 1)

        # summary of the day
        style_row(row, dheader_style)
        ws.merge_cells("{ca}{r}:{cb}{r}".format(
            r=row,
            ca=column_to_letter(transcript_col),
            cb=column_to_letter(location_col)))
        # header
        for colnum, header in enumerate(day_headers):
            std_write(row, colnum + 1, header, dheader_style)
        row += 1

        range_state = "{sl}{rs}:{sl}{re}".format(
            sl=state_letter, rs=day_start, re=day_end)
        range_issue = "{il}{rs}:{il}{re}".format(
            il=issue_letter, rs=day_start, re=day_end)

        # one line per reason
        for reason in reasons:
            style_row(row, std_style)
            apply_first_cols_style(row, 2, dheader_style)
            ws.merge_cells("{ca}{r}:{cb}{r}".format(
                r=row,
                ca=column_to_letter(transcript_col),
                cb=column_to_letter(location_col)))
            std_write(row, summary_dcol, reason, std_style)

            # formula for totals
            borno_cell = ws.cell(row=row, column=borno_dcol)
            borno_cell.set_explicit_value(
                value='=COUNTIFS({rs},"{n}",{ri},"{i}")'.format(
                    ri=range_issue, rs=range_state,
                    i=reason, n=day_headers[borno_dcol - 1]),
                data_type=borno_cell.TYPE_FORMULA)
            apply_style(borno_cell, value_style)

            adamawa_cell = ws.cell(row=row, column=adamawa_dcol)
            adamawa_cell.set_explicit_value(
                value='=COUNTIFS({rs},"{n}",{ri},"{i}")'.format(
                    ri=range_issue, rs=range_state,
                    i=reason, n=day_headers[adamawa_dcol - 1]),
                data_type=adamawa_cell.TYPE_FORMULA)
            apply_style(adamawa_cell, value_style)

            yobe_cell = ws.cell(row=row, column=yobe_dcol)
            yobe_cell.set_explicit_value(
                value='=COUNTIFS({rs},"{n}",{ri},"{i}")'.format(
                    ri=range_issue, rs=range_state,
                    i=reason, n=day_headers[yobe_dcol - 1]),
                data_type=yobe_cell.TYPE_FORMULA)
            apply_style(yobe_cell, value_style)

            total_cell = ws.cell(row=row, column=total_dcol)
            total_cell.set_explicit_value(
                value='=SUM({al}{r}:{yl}{r})'.format(
                    al=adamawa_letter, r=row, yl=yobe_letter),
                data_type=total_cell.TYPE_FORMULA)
            apply_style(total_cell, value_style)

            row += 1

        range_reasons = day_end, day_end + len(reasons)

        # total line for day
        style_row(row, bold_style)
        apply_first_cols_style(row, 2, dheader_style)
        ws.merge_cells("{ca}{r}:{cb}{r}".format(
            r=row,
            ca=column_to_letter(transcript_col),
            cb=column_to_letter(location_col)))
        std_write(row, summary_dcol, "Total", bold_style)

        # formula for totals
        borno_cell = ws.cell(row=row, column=borno_dcol)
        borno_cell.set_explicit_value(
            value='=SUM({})'.format(
                "{sl}{rs}:{sl}{re}".format(sl=borno_letter,
                                           rs=range_reasons[0],
                                           re=range_reasons[1])),
            data_type=borno_cell.TYPE_FORMULA)
        apply_style(borno_cell, value_style)

        adamawa_cell = ws.cell(row=row, column=adamawa_dcol)
        adamawa_cell.set_explicit_value(
            value='=SUM({})'.format(
                "{sl}{rs}:{sl}{re}".format(sl=adamawa_letter,
                                           rs=range_reasons[0],
                                           re=range_reasons[1])),
            data_type=borno_cell.TYPE_FORMULA)
        apply_style(adamawa_cell, value_style)

        yobe_cell = ws.cell(row=row, column=yobe_dcol)
        yobe_cell.set_explicit_value(
            value='=SUM({})'.format(
                "{sl}{rs}:{sl}{re}".format(sl=yobe_letter,
                                           rs=range_reasons[0],
                                           re=range_reasons[1])),
            data_type=yobe_cell.TYPE_FORMULA)
        apply_style(yobe_cell, value_style)

        std_write(row, total_dcol, day_requests.count(), value_style)
        row += 1

    week_end = copy.copy(row - 1)

    # summary of the week
    style_row(row, wheader_style)
    ws.merge_cells("{ca}{r}:{cb}{r}".format(
        r=row,
        ca=column_to_letter(transcript_col),
        cb=column_to_letter(location_col)))
    # header
    for colnum, header in enumerate(day_headers):
        std_write(row, colnum + 1, header, wheader_style)
    std_write(row, summary_dcol, "Summary for the week", wheader_style)
    row += 1

    range_state = "{sl}{rs}:{sl}{re}".format(
        sl=state_letter, rs=data_first_row, re=week_end)
    range_issue = "{il}{rs}:{il}{re}".format(
        il=issue_letter, rs=data_first_row, re=week_end)

    # one line per reason
    for reason in reasons:
        style_row(row, std_style)
        apply_first_cols_style(row, 2, wheader_style)
        ws.merge_cells("{ca}{r}:{cb}{r}".format(
            r=row,
            ca=column_to_letter(transcript_col),
            cb=column_to_letter(location_col)))
        std_write(row, summary_dcol, reason, std_style)

        # formula for totals
        borno_cell = ws.cell(row=row, column=borno_dcol)
        borno_cell.set_explicit_value(
            value='=COUNTIFS({rs},"{n}",{ri},"{i}")'.format(
                ri=range_issue, rs=range_state,
                i=reason, n=day_headers[borno_dcol - 1]),
            data_type=borno_cell.TYPE_FORMULA)
        apply_style(borno_cell, value_style)

        adamawa_cell = ws.cell(row=row, column=adamawa_dcol)
        adamawa_cell.set_explicit_value(
            value='=COUNTIFS({rs},"{n}",{ri},"{i}")'.format(
                ri=range_issue, rs=range_state,
                i=reason, n=day_headers[adamawa_dcol - 1]),
            data_type=adamawa_cell.TYPE_FORMULA)
        apply_style(adamawa_cell, value_style)

        yobe_cell = ws.cell(row=row, column=yobe_dcol)
        yobe_cell.set_explicit_value(
            value='=COUNTIFS({rs},"{n}",{ri},"{i}")'.format(
                ri=range_issue, rs=range_state,
                i=reason, n=day_headers[yobe_dcol - 1]),
            data_type=yobe_cell.TYPE_FORMULA)
        apply_style(yobe_cell, value_style)

        total_cell = ws.cell(row=row, column=total_dcol)
        total_cell.set_explicit_value(
            value='=SUM({al}{r}:{yl}{r})'.format(
                al=adamawa_letter, r=row, yl=yobe_letter),
            data_type=total_cell.TYPE_FORMULA)
        apply_style(total_cell, value_style)

        row += 1

    range_reasons = week_end, week_end + len(reasons)

    # total line for week
    style_row(row, bold_style)
    apply_first_cols_style(row, 2, wheader_style)
    ws.merge_cells("{ca}{r}:{cb}{r}".format(
        r=row,
        ca=column_to_letter(transcript_col),
        cb=column_to_letter(location_col)))
    std_write(row, summary_dcol, "Total", bold_style)

    # formula for totals
    borno_cell = ws.cell(row=row, column=borno_dcol)
    borno_cell.set_explicit_value(
        value='=SUM({})'.format(
            "{sl}{rs}:{sl}{re}".format(sl=borno_letter,
                                       rs=range_reasons[0],
                                       re=range_reasons[1])),
        data_type=borno_cell.TYPE_FORMULA)
    apply_style(borno_cell, value_style)

    adamawa_cell = ws.cell(row=row, column=adamawa_dcol)
    adamawa_cell.set_explicit_value(
        value='=SUM({})'.format(
            "{sl}{rs}:{sl}{re}".format(sl=adamawa_letter,
                                       rs=range_reasons[0],
                                       re=range_reasons[1])),
        data_type=adamawa_cell.TYPE_FORMULA)
    apply_style(adamawa_cell, value_style)

    yobe_cell = ws.cell(row=row, column=yobe_dcol)
    yobe_cell.set_explicit_value(
        value='=SUM({})'.format(
            "{sl}{rs}:{sl}{re}".format(sl=yobe_letter,
                                       rs=range_reasons[0],
                                       re=range_reasons[1])),
        data_type=yobe_cell.TYPE_FORMULA)
    apply_style(yobe_cell, value_style)

    std_write(row, total_dcol, hr_requests.count(), value_style)
    row += 1

    stream = io.BytesIO()
    wb.save(stream)
    return stream


@login_required
def reports(request):
    now = datetime.datetime.now()
    start = now - datetime.timedelta(days=31)
    context = {
        'weeks': weeks_for_period(start, now),
        'clusters': Cluster.objects.exclude(
            slug__in=('foreign', 'other', 'english'))
    }
    return render(request, 'reports.html', context)


@login_required
def report(request, lang, period):

    file_name = "logbook-{p}-{l}.xlsx".format(p=period, l=lang)
    file_content = gen_report(lang, period)
    file_content = file_content.getvalue()

    response = HttpResponse(
        file_content,
        content_type='application/vnd.openxmlformats-officedocument.'
                     'spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s"' % file_name
    response['Content-Length'] = len(file_content)

    return response
